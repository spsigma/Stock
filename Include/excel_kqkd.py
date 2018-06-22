from openpyxl import load_workbook
import os

os.getcwd()
os.chdir('C:\\Users\SP\Desktop')
wb = load_workbook(filename='kqkd.xlsx')
ws = wb.active

print('Nhap ma co phieu: ', end=' ')
macp = input()
macp = macp.upper()


for i in range(732):
    if (ws.cell(row=i+7, column=3).value) == macp:
        cp = macp
        print(cp, end=  '  ')
        growth = ws.cell(row = i+7, column = 8).value
        print( growth)