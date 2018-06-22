from openpyxl import load_workbook
import os

os.getcwd()
os.chdir('C:\\Users\SP\Desktop')
wb = load_workbook(filename='kqkd.xlsx')
ws = wb['TA']

def support(cp):
    for i in range(10):
        if (ws.cell(row=i+3, column=1).value) == cp:
            s1 = ws.cell(row=i+3, column=2).value
            s2 = ws.cell(row=i+3, column=3).value
            s3 = ws.cell(row=i+3, column=4).value

    print(s1, '  ', s2, '   ', s3)





