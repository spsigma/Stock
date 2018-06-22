import os
import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from tabulate import tabulate

os.getcwd()
os.chdir('C:\\Users\SP\Desktop')
wb = load_workbook(filename='kqkd.xlsx')
ws = wb['kqkd']

def tangtruong(cophieu):
    for i in range(732):
        if (ws.cell(row=i + 7, column=3).value) == cophieu:
            growth = ws.cell(row=i + 7, column=9).value

    return growth

def tangtruong_quytruoc(cophieu):
    for i in range(732):
        if (ws.cell(row=i + 7, column=3).value) == cophieu:
            growth_quytruoc = ws.cell(row=i + 7, column=8).value

    return growth_quytruoc


def codong(cp):
    page = requests.get('http://www.cophieu68.vn/shareholder.php?id=' + cp)
    soup = BeautifulSoup(page.content, 'html.parser')

    data = []
    table = soup.findAll('table')
    table = table[2]
    rows = table.findAll('tr')

    df = pd.read_html(str(table))
    print(tabulate(df[0], headers='keys', tablefmt='psql'))

def support(cp):
    os.chdir('C:\\Users\SP\Desktop')
    wb = load_workbook(filename='kqkd.xlsx')
    ws = wb['TA']
    for i in range(10):
        if (ws.cell(row=i + 3, column=1).value) == cp:
            s1 = ws.cell(row=i + 3, column=2).value
            s2 = ws.cell(row=i + 3, column=3).value
            s3 = ws.cell(row=i + 3, column=4).value

    print(s1, '  ', s2, '   ', s3)

def resistence(cp):
    os.chdir('C:\\Users\SP\Desktop')
    wb = load_workbook(filename='kqkd.xlsx')
    ws = wb['TA']
    for i in range(10):
        if (ws.cell(row=i + 3, column=1).value) == cp:
            r1 = ws.cell(row=i + 3, column=5).value
            r2 = ws.cell(row=i + 3, column=6).value
            r3 = ws.cell(row=i + 3, column=7).value

    print(r1, '  ', r2, '   ', r3)


while True:
    print('Nhập cổ phiếu : ', end='')
    cp=input()
    cp = cp.upper()
    page=requests.get('http://www.cophieu68.vn/snapshot.php?id='+cp)
    soup=BeautifulSoup(page.content, 'html.parser')
    price: object=soup.find('strong', {'id':"stockname_close"})
    price=price.text
    price=float(price)

    print('Giá cổ phiếu         : ',price)

    data=[]
    table=soup.find('div', {'id':'snapshot_trading'})
    rows=table.findAll('tr')

    for x, row in enumerate(rows[1:]):
        cols=row.findAll('td')
        for y,col in enumerate(cols):
            data.append([])
            data[x].append(col)


    beta=(data[13][1])
    beta=beta.text
    roe=(data[12][1])
    roe=roe.text
    pe=(data[8][1])
    pe=pe.text
    pe=pe.replace('lần','')
    pe=float(pe)
    mc=(data[9][1])
    mc=mc.text
    mc=mc.replace('Tỷ','')
    mcfloat=mc.replace(',','')
    mcfloat=float(mcfloat)
    shares=(data[10][1])
    shares=shares.text
    shares=shares.replace('triệu', '')
    sharesfloat=shares.replace(',', '')
    sharesfloat=float(sharesfloat)
    bv=(data[11][1])
    bv=bv.text
    bv=bv.replace('ngàn','')
    bv=float(bv)
    kltb = data[4][1]



    print('Beta của cổ phiếu    : ',beta)
    print('ROE của cổ phiếu     : ',roe)
    print('PE của cổ phiếu      : ',pe)
    print('Market Cap           : ',mc)
    print('Lợi nhuận 4 quý      : ', round(mcfloat/pe,2))
    print('Cổ phiếu lưu hành    : ', shares)
    print('Giá trị sổ sách      : ', bv)
    print('Số lượng GDTB        : ', kltb.text)
    print('\n')

    print('Xem các ngưỡng kỹ thuật giá cổ phiếu ?', end='')
    traloi = input()
    if traloi == 'y':
        print('Các ngưỡng hỗ trợ    : ', end= '')
        support(cp)
        print('Các ngưỡng cản       : ', end ='')
        resistence(cp)


    print('Kiểm tra tăng trưởng của DN ?', end = ' ')
    kiemtra = input()
    if kiemtra=='y':
        print('Tăng trưởng vs. quý 4: ', tangtruong_quytruoc(cp))
        print('Tăng trưởng quý 1    : ', tangtruong(cp))
        g = tangtruong(cp)
        if (g<0):
            print('Chú ý DN đang tăng trưởng âm, phải có event-driven để đầu tư')

        if (0<g<10):
            print('Cổ phiếu này không phải cổ phiếu tăng trưởng')
            print('PEG = ', (pe/g))
        elif (10<g<20):
            print('Tăng trưởng mức độ thấp')
            print('PEG có thể hấp dẫn: ', (pe/g))
        elif (g>20):
            print('Cổ phiếu tăng trưởng cao nên chú ý các yếu tố động cơ tăng trưởng mạnh mẽ')
            print('PEG rất hấp dẫn: ', (pe/g))

    print('Có xem danh sách cổ đông không ?, ', end= '  ')
    xem = input()
    if xem=='y':
        codong(cp)

    print('Tiếp tục ? : ', end='')
    key=input()
    if key=='n':
      break